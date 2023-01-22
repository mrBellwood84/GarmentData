import os
from openpyxl import Workbook
from modules.config import Config
from modules.tableData import TableData

class ExcelGenerator():

    def __init__(self, data: TableData, config: Config) -> None:
        self.data = data
        self.config = config
        
        self.excelFile = self.__getPath()

    def __getPath(self):
        path = self.config.outputFolder if self.config.outputFolder else os.path.join(os.getcwd(), "Output")
        filename = f"{self.config.outputFileName}.xlsx" if self.config.outputFileName else "result.xlsx"
        return os.path.join(path, filename)

    def create(self):
        wb = Workbook()
        ws = wb.create_sheet(self.data.tableName)

        for rowIndex, row in enumerate(self.data.tableData):
            for colIndex, cellData in enumerate(row):
                ws.cell(row=rowIndex+1, column=colIndex+1, value=cellData)

        wb.save(self.excelFile)
    
    



    
